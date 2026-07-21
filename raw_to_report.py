# -*- coding: utf-8 -*-
"""
로우데이터(일별 원본 데이터) -> 주간 리포트 시트 자동 생성기.

지원 시트: NAVER GFA_Conf, NAVER GFA_Pet, NAVER BS_Pet
(NAVER BS_Conf/SP_Conf/KAKAO_DA는 주간 광고비가 로우데이터에 없고 매주 금액이 일정하지 않아 제외)

로우데이터 컬럼(고정 포맷): Date, Media, Device, 캠페인, 그룹, 노출, 클릭, 광고비, 구매완료수, 매출액

생성된 시트는 app.py의 CELL_MAP과 동일한 행 구조로 [Comment] 자리를 비워두고,
그 아래 Weekly 성과 표를 만든다 -> 기존 브리핑/코멘트 생성 로직을 그대로 이어서 쓸 수 있다.
"""
import re
from datetime import date, timedelta
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 브랜드검색(BS) 시트는 로우데이터에 실제 광고비가 없어 월예산을 주간으로 환산해서 사용
# (검증됨: 주간 광고비 = 월예산 x 7/30, 실제 리포트값과 정확히 일치)
BS_PET_MONTHLY_BUDGET = {
    '그리니즈': 800000, '시저': 1700000, '템테이션': 800000, '쉬바': 800000, '위스카스': 800000,
}
BS_PET_BRANDS = list(BS_PET_MONTHLY_BUDGET.keys())

HEADERS = ['WEEK', 'IMPS', 'CLICK', 'CTR', 'CPC', 'COST (-vat)',
           'TRANS.\n(구매완료 수)', 'SALES\n(구매완료 매출액)', 'CPA', 'CVR', 'ROAS']

# 각 시트의 코멘트 예약 영역 (app.py의 CELL_MAP과 동일한 행 번호를 써야 함)
COMMENT_LAYOUT = {
    'NAVER GFA_Conf': [(5, '[Comment]'), (9, '<ADVoost>'), (13, '<NDA>'), (17, '<쇼핑프로모션>')],
    'NAVER GFA_Pet': [(5, '[Comment]'), (9, '<ADVoost>'), (12, '<NDA>'), (15, '<쇼핑프로모션>'), (20, '<스마트채널>')],
    'NAVER BS_Pet': [(5, '[Comment]')],
}
COMMENT_BLOCK_END = {'NAVER GFA_Conf': 22, 'NAVER GFA_Pet': 24, 'NAVER BS_Pet': 19}

SHEET_GROUP_ORDER = {
    'NAVER GFA_Conf': ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션'],
    'NAVER GFA_Pet': ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션', '스마트채널'],
    'NAVER BS_Pet': ['TOTAL'] + BS_PET_BRANDS,
}

# 리포트의 COST(-vat)는 부가세 제외 금액. GFA 로우데이터의 광고비는 부가세 포함 금액이라
# /1.1 보정이 필요함 (실제 완성 리포트 수치와 대조 검증: 전 채널에서 정확히 1.1배 차이 확인됨)
GFA_VAT_DIVISOR = 1.1

WEEK_RE = re.compile(r'^\d+월\s*\d+주차\s*\(')


def monday_of(d):
    return d - timedelta(days=d.weekday())


def week_index_in_month(monday):
    idx = 1
    cur = monday - timedelta(days=7)
    while cur.month == monday.month:
        idx += 1
        cur -= timedelta(days=7)
    return idx


def week_label(d):
    ws = monday_of(d)
    we = ws + timedelta(days=6)
    idx = week_index_in_month(ws)
    label = "{}월 {}주차 ({}/{}~{}/{})".format(ws.month, idx, ws.month, ws.day, we.month, we.day)
    return label, ws


def to_date(v):
    if isinstance(v, str):
        return date(*map(int, v.split('-')))
    return v.date() if hasattr(v, 'date') else v


def classify_gfa(campaign):
    if campaign.startswith('(Mars Conf)'):
        sheet = 'NAVER GFA_Conf'
    elif campaign.startswith('(Mars Pet)'):
        sheet = 'NAVER GFA_Pet'
    else:
        return None
    name = campaign.split(')', 1)[1].strip()
    if 'ADVoost' in name:
        ch = 'ADVoost'
    elif 'NDA' in name:
        ch = 'NDA'
    elif '스마트채널' in name:
        ch = '스마트채널'
    elif '쇼핑프로모션' in name:
        ch = '쇼핑프로모션'
    else:
        ch = name
    return sheet, ch


def classify_bs(campaign, group):
    if 'PET' in campaign and '브랜드검색' in campaign:
        for b in BS_PET_BRANDS:
            if group and b in group:
                return 'NAVER BS_Pet', b
    return None  # Conf 브랜드검색은 이번 자동화 대상 아님


def parse_raw(wb_raw):
    ws = wb_raw['Msrs Daily'] if 'Msrs Daily' in wb_raw.sheetnames else wb_raw[wb_raw.sheetnames[0]]
    agg = defaultdict(lambda: defaultdict(lambda: [0, 0, 0.0, 0, 0.0]))
    max_date = None
    min_date = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        d, media, device, camp, group, imps, clicks, cost, trans, sales = row[:10]
        if d is None or camp is None:
            continue
        dd = to_date(d)
        if max_date is None or dd > max_date:
            max_date = dd
        if min_date is None or dd < min_date:
            min_date = dd
        label, wmon = week_label(dd)

        classified = None
        if media == 'GFA':
            classified = classify_gfa(camp)
        elif media == 'NaverBS':
            classified = classify_bs(camp, group)
        if not classified:
            continue
        sheet, key = classified

        cost_val = cost or 0
        if media == 'GFA':
            cost_val = cost_val / GFA_VAT_DIVISOR

        bucket = agg[(sheet, key)][(label, wmon)]
        bucket[0] += imps or 0
        bucket[1] += clicks or 0
        bucket[2] += cost_val
        bucket[3] += trans or 0
        bucket[4] += sales or 0

    # 로우데이터 범위의 앞/뒤로 걸쳐 있어 '일부 날짜만' 잡힌 주는 완전하지 않으므로 0으로 비워둔다.
    # (뒤쪽: 아직 안 끝난 진행중인 주 / 앞쪽: 로우데이터가 그 주 중간부터 시작해서 앞부분이 없는 주)
    # 이렇게 해야 이미 완성되어 있던 과거 주 값을 불완전한 값으로 잘못 덮어쓰지 않는다.
    if max_date is not None and min_date is not None:
        for key_, weeks in agg.items():
            for (label, wmon), vals in weeks.items():
                week_end = wmon + timedelta(days=6)
                if week_end > max_date or wmon < min_date:
                    for i in range(5):
                        vals[i] = 0
    return agg


def compute_row(imps, clicks, cost, trans, sales):
    ctr = clicks / imps if imps else 0
    cpc = cost / clicks if clicks else 0
    cpa = cost / trans if trans else 0
    cvr = trans / clicks if clicks else 0
    roas = sales / cost if cost else 0
    return [imps, clicks, ctr, cpc, cost, trans, sales, cpa, cvr, roas]


def build_report(raw_bytes):
    """로우데이터 바이트를 받아 새 리포트 워크북(openpyxl Workbook)을 반환한다."""
    wb_raw = openpyxl.load_workbook(raw_bytes, data_only=True)
    agg = parse_raw(wb_raw)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # BS_Pet 브랜드는 실제 광고비 대신 월예산 환산값을 쓰므로, TOTAL 합산 전에 미리 대체해둔다
    # (단, 아직 진행중이라 0으로 비워진 주는 예산도 채우지 않고 그대로 0 유지)
    for group in BS_PET_BRANDS:
        weeks = agg.get(('NAVER BS_Pet', group))
        if not weeks:
            continue
        budget_cost = BS_PET_MONTHLY_BUDGET.get(group, 0) * 7 / 30
        for wk, vals in weeks.items():
            is_incomplete = vals[0] == 0 and vals[1] == 0 and vals[3] == 0 and vals[4] == 0
            vals[2] = 0 if is_incomplete else budget_cost

    # TOTAL(시트 합계)을 위해 그룹별 원자료를 합산 (위에서 보정된 값 기준)
    per_sheet_week_totals = defaultdict(lambda: defaultdict(lambda: [0, 0, 0.0, 0, 0.0]))
    for (sheet, key), weeks in agg.items():
        for wk, vals in weeks.items():
            t = per_sheet_week_totals[sheet][wk]
            for i in range(5):
                t[i] += vals[i]

    for sheet_name, groups in SHEET_GROUP_ORDER.items():
        ws = wb.create_sheet(sheet_name)
        ws['B3'] = sheet_name
        ws['B3'].font = Font(bold=True, size=12)

        for row_i, label in COMMENT_LAYOUT[sheet_name]:
            c = ws.cell(row=row_i, column=2, value=label)
            c.font = Font(bold=True)

        r = COMMENT_BLOCK_END[sheet_name] + 2
        ws.cell(row=r, column=2, value='Weekly 성과').font = Font(bold=True)
        r += 2

        for group in groups:
            ws.cell(row=r, column=3, value=group).font = Font(bold=True)
            r += 1
            for c, h in enumerate(HEADERS, start=2):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='F2F2F2')
            r += 1

            if group == 'TOTAL':
                weeks_data = per_sheet_week_totals[sheet_name]
            else:
                weeks_data = agg.get((sheet_name, group), {})

            weeks_sorted = sorted(weeks_data.items(), key=lambda kv: kv[0][1])
            first_data_row = r
            for (label, wmon), vals in weeks_sorted:
                if group == 'TOTAL':
                    imps, clicks, cost, trans, sales = vals
                elif group in BS_PET_BRANDS:
                    imps, clicks, cost, trans, sales = vals  # cost는 위에서 이미 예산으로 대체됨
                else:
                    imps, clicks, cost, trans, sales = vals
                computed = compute_row(imps, clicks, cost, trans, sales)
                ws.cell(row=r, column=2, value=label)
                for c, v in enumerate(computed, start=3):
                    ws.cell(row=r, column=c, value=round(v, 6) if isinstance(v, float) else v)
                r += 1
            last_data_row = r - 1

            if last_data_row >= first_data_row + 1:
                ws.cell(row=r, column=2, value='WoW').font = Font(italic=True)
                for c in range(3, 13):
                    col_letter = get_column_letter(c)
                    prev_cell = "{}{}".format(col_letter, last_data_row - 1)
                    curr_cell = "{}{}".format(col_letter, last_data_row)
                    formula = "=IFERROR(({}-{})/{},\"\")".format(curr_cell, prev_cell, prev_cell)
                    ws.cell(row=r, column=c, value=formula)
                    ws.cell(row=r, column=c).number_format = '0.0%'
                r += 1
            r += 1

        widths = [22, 12, 12, 10, 10, 14, 10, 14, 10, 10, 10]
        for col, width in zip('BCDEFGHIJKL', widths):
            ws.column_dimensions[col].width = width

    wb.calculation.fullCalcOnLoad = True
    return wb


# ─────────────────────────────────────────────────────────────
# 디자인을 그대로 보존하는 업데이트 방식:
# "지난 주에 쓰던 실제 파일"을 그대로 열어서, 이번 주 해당하는 행의 숫자만 갱신한다.
# (서식·다른 시트·차트 등은 원본 파일 그대로 유지됨 — 새로 만드는 게 아니라 "고쳐 쓰는" 방식)
# ─────────────────────────────────────────────────────────────
SHEET_GROUP_KEYWORDS = {
    'NAVER GFA_Conf': ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션'],
    'NAVER GFA_Pet':  ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션', '스마트채널'],
    'NAVER BS_Pet':   ['TOTAL', '그리니즈', '시저', '템테이션', '쉬바', '위스카스'],
}


def _nearest_group_label(ws, row):
    bad = {'WEEK', 'IMPS', 'CLICK', 'CTR', 'CPC', 'COST (-vat)',
           'TRANS.\n(구매완료 수)', 'SALES\n(구매완료 매출액)', 'CPA', 'CVR',
           'ROAS', 'WoW', '<그룹 별 상세 성과>', None}
    for rr in range(row, max(row - 250, 0), -1):
        for col in (3, 2):
            v = ws.cell(row=rr, column=col).value
            if (isinstance(v, str) and v not in bad and not WEEK_RE.match(v)
                    and not v.startswith('=') and v.strip() != ''):
                return v
    return None


def _find_group_tables_in_workbook_sheet(ws):
    tables = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and WEEK_RE.match(cell.value):
                r = cell.row
                gname = _nearest_group_label(ws, r - 2)
                tables.setdefault(gname, []).append(r)
    for g in tables:
        tables[g].sort()
    return tables


def update_existing_report(prev_report_bytes, raw_bytes):
    """지난 주 실제 파일(서식 그대로) + 이번 로우데이터 -> 이번 주 숫자만 갱신된 워크북.

    - 완료된 주(직전 raw 데이터 기준)에 해당하는 행을 찾아 그 행의 수치만 실제 계산값으로 덮어쓴다.
    - 서식, 다른 시트, 아직 안 온 미래 주(빈 값/수식)는 전혀 건드리지 않는다.
    """
    wb = openpyxl.load_workbook(prev_report_bytes, data_only=False)
    agg = parse_raw(openpyxl.load_workbook(raw_bytes, data_only=True))

    # BS_Pet 브랜드 예산 반영 + 진행중인 주 0 처리 (build_report와 동일 규칙)
    for group in BS_PET_BRANDS:
        weeks = agg.get(('NAVER BS_Pet', group))
        if not weeks:
            continue
        budget_cost = BS_PET_MONTHLY_BUDGET.get(group, 0) * 7 / 30
        for wk, vals in weeks.items():
            is_incomplete = vals[0] == 0 and vals[1] == 0 and vals[3] == 0 and vals[4] == 0
            vals[2] = 0 if is_incomplete else budget_cost

    per_sheet_week_totals = defaultdict(lambda: defaultdict(lambda: [0, 0, 0.0, 0, 0.0]))
    for (sheet, key), weeks in agg.items():
        for wk, vals in weeks.items():
            t = per_sheet_week_totals[sheet][wk]
            for i in range(5):
                t[i] += vals[i]

    updated_cells = []
    for sheet_name, keywords in SHEET_GROUP_KEYWORDS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tables = _find_group_tables_in_workbook_sheet(ws)

        for gname, rows in tables.items():
            if gname is None:
                continue
            matched_kw = next((kw for kw in keywords if kw in gname), None)
            if not matched_kw:
                continue

            if matched_kw == 'TOTAL':
                weeks_data = per_sheet_week_totals.get(sheet_name, {})
            else:
                weeks_data = agg.get((sheet_name, matched_kw), {})

            for (label, wmon), vals in weeks_data.items():
                sales = vals[4] if matched_kw != 'TOTAL' else vals[4]
                # 아직 진행중(0)인 주는 건드리지 않음 -> 기존 서식/수식 그대로 유지
                if all(v == 0 for v in vals):
                    continue
                # 이 그룹의 표에서 해당 주(label)에 맞는 행을 찾는다
                target_row = None
                for r in rows:
                    if ws.cell(row=r, column=2).value == label:
                        target_row = r
                        break
                if target_row is None:
                    continue  # 표에 아직 이 주가 준비되어 있지 않음 (템플릿을 더 늘려야 함)

                imps, clicks, cost, trans, sales = vals
                computed = compute_row(imps, clicks, cost, trans, sales)
                for c, v in enumerate(computed, start=3):
                    ws.cell(row=target_row, column=c, value=round(v, 6) if isinstance(v, float) else v)
                updated_cells.append((sheet_name, matched_kw, label, target_row))

    wb.calculation.fullCalcOnLoad = True
    return wb, updated_cells
