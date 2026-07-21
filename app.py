# -*- coding: utf-8 -*-
"""
주간 광고 리포트 코멘트 자동화 앱 (Streamlit) — 완전 무료 버전 (API 키 불필요)

로우데이터 또는 완성된 리포트 파일(.xlsx)을 업로드하면:
  1) 시트별 브랜드/채널 WEEK 테이블에서 최신 주 vs 직전 주 수치를 스스로 찾아 증감률을 계산하고
  2) 화면에서 이번 주 특이사항(타임딜, 브랜드데이 등)을 입력받아
  3) 증감률 크기에 따라 미리 정해둔 문구 패턴을 조합해 코멘트 문장을 만든 뒤
  4) 화면에서 검수/수정하고
  5) 실제 셀에 반영한 엑셀 파일을 다운로드한다.

추가로, 완성된 리포트 파일을 올리면 시트/채널/브랜드별 주차 추이 그래프와
비교 그래프, 상세 데이터 표를 보여주는 성과 대시보드도 제공한다.

Anthropic API 키가 필요 없는 완전 무료 버전. 문장은 규칙 기반이라 다소 정형화되어 있음 —
필요하면 화면에서 직접 다듬을 수 있다.
"""
import io
import re
from datetime import datetime

import streamlit as st
import openpyxl
import pandas as pd
import plotly.graph_objects as go

import raw_to_report

# ─────────────────────────────────────────────────────────────
# 1) 시트 구조 설정
# ─────────────────────────────────────────────────────────────
SHEET_GROUP_KEYWORDS = {
    'NAVER GFA_Conf': ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션'],
    'NAVER GFA_Pet':  ['TOTAL', 'ADVoost', 'NDA', '쇼핑프로모션', '스마트채널'],
    'NAVER BS_Pet':   ['TOTAL', '그리니즈', '시저', '템테이션', '쉬바', '위스카스'],
}

CELL_MAP = {
    'NAVER GFA_Conf': {
        'TOTAL': ['B6', 'B7'], 'ADVoost': ['B10', 'B11'], 'NDA': ['B14', 'B15'],
        '쇼핑프로모션': ['B18', 'B19', 'B20'],
    },
    'NAVER GFA_Pet': {
        'TOTAL': ['B6', 'B7'], 'ADVoost': ['B10'], 'NDA': ['B13'],
        '쇼핑프로모션': ['B16', 'B17', 'B18'], '스마트채널': ['B21', 'B22'],
    },
    'NAVER BS_Pet': {
        'TOTAL': ['B6', 'B7'], '그리니즈': ['B9'], '시저': ['B11'],
        '템테이션': ['B13'], '쉬바': ['B15'], '위스카스': ['B17'],
    },
}

BS_PET_BRAND_NUMBER = {'그리니즈': 1, '시저': 2, '템테이션': 3, '쉬바': 4, '위스카스': 5}

WEEK_RE = re.compile(r'^\d+월\s*\d+주차\s*\(')
BAD_LABELS = {'WEEK', 'IMPS', 'CLICK', 'CTR', 'CPC', 'COST (-vat)',
              'TRANS.\n(구매완료 수)', 'SALES\n(구매완료 매출액)', 'CPA', 'CVR',
              'ROAS', 'WoW', '<그룹 별 상세 성과>', None}
COLS = ['IMPS', 'CLICK', 'CTR', 'CPC', 'COST', 'TRANS', 'SALES', 'CPA', 'CVR', 'ROAS']


# ─────────────────────────────────────────────────────────────
# 2) 데이터 추출
# ─────────────────────────────────────────────────────────────
def nearest_group_label(ws, row):
    for rr in range(row, max(row - 40, 0), -1):
        for col in (3, 2):
            v = ws.cell(row=rr, column=col).value
            if isinstance(v, str) and v not in BAD_LABELS and not WEEK_RE.match(v) and v.strip() != '':
                return v
    return None


def find_group_tables(ws):
    tables = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and WEEK_RE.match(cell.value):
                r = cell.row
                header_row = r - 1
                gname = nearest_group_label(ws, header_row - 1)
                vals = [ws.cell(row=r, column=c).value for c in range(3, 13)]
                tables.setdefault(gname, []).append((r, cell.value, vals))
    for g in tables:
        tables[g].sort(key=lambda x: x[0])
    return tables


def latest_complete_pair(rows):
    last_idx = None
    for i, (r, label, vals) in enumerate(rows):
        sales = vals[6] if len(vals) > 6 else None
        if sales not in (None, 0):
            last_idx = i
    if last_idx is None or last_idx == 0:
        return None
    return rows[last_idx - 1], rows[last_idx]


def pct(a, b):
    if a in (0, None) or b is None:
        return None
    return round((b - a) / a * 100, 1)


def build_brief(wb):
    brief = {}
    for sheet_name, group_keywords in SHEET_GROUP_KEYWORDS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tables = find_group_tables(ws)
        sheet_brief = {}
        for gname, rows in tables.items():
            if gname is None:
                continue
            matched_kw = next((kw for kw in group_keywords if kw in gname), None)
            if not matched_kw or matched_kw in sheet_brief:
                continue
            pair = latest_complete_pair(rows)
            if not pair:
                continue
            (r1, label1, v1), (r2, label2, v2) = pair
            deltas = {col: {'prev': a, 'curr': b, 'pct_change': pct(a, b)}
                      for col, a, b in zip(COLS, v1, v2)}
            sheet_brief[matched_kw] = {
                'prev_week': label1, 'curr_week': label2,
                'metrics': deltas, 'user_note': '',
            }
        if sheet_brief:
            brief[sheet_name] = sheet_brief
    return brief


def build_full_series(wb):
    """대시보드용: 시트/채널·브랜드별 '전체' 주차 시계열 데이터를 추출한다.

    build_brief는 최신 2개 주만 비교하지만, 대시보드는 추이 그래프를 그려야 하므로
    데이터가 채워진 모든 주차를 가져온다. (진행 중이라 비어있는 주는 제외)
    """
    series = {}
    for sheet_name, keywords in SHEET_GROUP_KEYWORDS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tables = find_group_tables(ws)
        sheet_series = {}
        for gname, rows in tables.items():
            if gname is None:
                continue
            matched_kw = next((kw for kw in keywords if kw in gname), None)
            if not matched_kw or matched_kw in sheet_series:
                continue
            pts = []
            for r, label, vals in rows:
                imps, clicks, ctr, cpc, cost, trans, sales, cpa, cvr, roas = vals
                if (imps in (None, 0)) and (sales in (None, 0)) and (cost in (None, 0)):
                    continue  # 진행 중이거나 비어있는 주는 제외
                pts.append({
                    'WEEK': label,
                    'IMPS': imps or 0, 'CLICK': clicks or 0, 'CTR': (ctr or 0) * 100,
                    'CPC': cpc or 0, 'COST': cost or 0, 'TRANS': trans or 0,
                    'SALES': sales or 0, 'CPA': cpa or 0, 'CVR': (cvr or 0) * 100,
                    'ROAS': (roas or 0) * 100,
                })
            if pts:
                sheet_series[matched_kw] = pts
        if sheet_series:
            series[sheet_name] = sheet_series
    return series


METRIC_LABELS = {
    'IMPS': 'IMPS (노출)', 'CLICK': 'CLICK (클릭)', 'CTR': 'CTR (%)', 'CPC': 'CPC (원)',
    'COST': 'COST (원, -vat)', 'TRANS': 'TRANS (구매완료수)', 'SALES': 'SALES (매출액, 원)',
    'CPA': 'CPA (원)', 'CVR': 'CVR (%)', 'ROAS': 'ROAS (%)',
}


# 주요지표(절대량 계열) = 막대, 보조지표(효율/비율 계열) = 선그래프
PRIMARY_METRICS = ['IMPS', 'CLICK', 'COST', 'TRANS', 'SALES']
SECONDARY_METRICS = ['CTR', 'CPC', 'CPA', 'CVR', 'ROAS']


def render_combo_chart(x_values, x_label, series_dict, primary_metrics, secondary_metrics, title=None):
    """주요지표는 막대(왼쪽 축), 보조지표는 선그래프(오른쪽 축)로 겹쳐 그린다."""
    fig = go.Figure()
    for m in primary_metrics:
        if m in series_dict:
            fig.add_bar(name=METRIC_LABELS[m], x=x_values, y=series_dict[m])
    for m in secondary_metrics:
        if m in series_dict:
            fig.add_trace(go.Scatter(
                name=METRIC_LABELS[m], x=x_values, y=series_dict[m],
                mode='lines+markers', yaxis='y2',
            ))
    layout = dict(
        barmode='group',
        bargap=0.35,        # 막대 그룹 사이 공백
        bargroupgap=0.2,    # 같은 그룹 내 막대 사이 공백 (막대가 너무 두꺼워지지 않도록)
        xaxis=dict(title=x_label),
        yaxis=dict(title='주요지표 (막대)'),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        margin=dict(t=70),
    )
    if secondary_metrics:
        layout['yaxis2'] = dict(title='보조지표 (선그래프, %/원)', overlaying='y', side='right')
    if title:
        layout['title'] = title
    fig.update_layout(**layout)
    st.plotly_chart(fig, use_container_width=True)


def infer_year_month_labels(week_labels):
    """주차 라벨(예: '7월 3주차 (7/15~7/21)')에는 연도가 없으므로, 라벨 순서(과거→최신)를 보고
    연도를 추정해 'YYYY-MM' 문자열 리스트로 반환한다.
    가장 최근 주는 오늘 날짜 기준 연도로 보고, 과거로 갈수록 월이 거꾸로 커지는 지점(연도 경계)에서
    연도를 하나씩 낮춘다."""
    months = [int(re.match(r'(\d+)월', w).group(1)) for w in week_labels]
    years = [None] * len(months)
    current_year = datetime.now().year
    prev_month = None
    for i in range(len(months) - 1, -1, -1):  # 최신 -> 과거 순으로 훑는다
        mo = months[i]
        if prev_month is not None and mo > prev_month:
            current_year -= 1
        years[i] = current_year
        prev_month = mo
    return [f"{y}-{m:02d}" for y, m in zip(years, months)]


TABLE_METRIC_ORDER = ['IMPS', 'CLICK', 'COST', 'TRANS', 'SALES', 'CTR', 'CPC', 'CVR', 'ROAS']
TABLE_METRIC_KR = {
    'IMPS': '노출수', 'CLICK': '클릭수', 'COST': '광고비', 'TRANS': '전환수', 'SALES': '매출',
    'CTR': 'CTR', 'CPC': 'CPC', 'CVR': 'CVR', 'ROAS': 'ROAS',
}


def format_value(metric, value):
    if metric in ('CTR', 'CVR'):
        return f"{value:.2f}%"
    if metric == 'ROAS':
        return f"{value:.0f}%"
    return f"{value:,.0f}"


def format_change(metric, curr, prev):
    """CTR/CVR은 이미 %로 표현된 지표이므로 %p(포인트) 차이로, 그 외 지표는 상대 증감률(%)로 표시."""
    if prev is None:
        return None
    if metric in ('CTR', 'CVR'):
        diff = curr - prev
        if abs(diff) < 0.05:
            return '—', '변동없음', '#888888'
        arrow = '▲' if diff > 0 else '▼'
        color = '#d93025' if diff > 0 else '#1a73e8'
        return arrow, f"{abs(diff):.1f}%p", color
    if prev == 0:
        return None
    pct_chg = (curr - prev) / prev * 100
    if abs(pct_chg) < 0.05:
        return '—', '변동없음', '#888888'
    arrow = '▲' if pct_chg > 0 else '▼'
    color = '#d93025' if pct_chg > 0 else '#1a73e8'
    return arrow, f"{abs(pct_chg):.1f}%", color


def render_group_table_html(groups_data, week_label):
    """선택한 주차 하나에 대해, 채널/브랜드별(+TOTAL) 지표 표를 HTML로 만든다.
    각 셀에는 값과 함께 직전 주 대비 증감(▲빨강/▼파랑)을 작게 표시한다."""
    table_groups = [g for g in groups_data if g != 'TOTAL']
    if 'TOTAL' in groups_data:
        table_groups.append('TOTAL')

    html = ['<table style="width:100%; border-collapse:collapse; font-size:13px;">']
    html.append('<thead><tr style="border-bottom:2px solid #333;">')
    html.append('<th style="text-align:left; padding:6px 10px;">영역</th>')
    for m in TABLE_METRIC_ORDER:
        html.append(f'<th style="text-align:right; padding:6px 10px;">{TABLE_METRIC_KR[m]}</th>')
    html.append('</tr></thead><tbody>')

    for g in table_groups:
        pts = groups_data[g]
        idx = next((i for i, p in enumerate(pts) if p['WEEK'] == week_label), None)
        if idx is None:
            continue
        curr = pts[idx]
        prev = pts[idx - 1] if idx > 0 else None
        is_total = (g == 'TOTAL')
        row_style = 'border-top:2px solid #333; font-weight:bold;' if is_total else 'border-bottom:1px solid #eee;'
        html.append(f'<tr style="{row_style}">')
        html.append(f'<td style="text-align:left; padding:6px 10px;">{g}</td>')
        for m in TABLE_METRIC_ORDER:
            val_str = format_value(m, curr[m])
            change = format_change(m, curr[m], prev[m] if prev else None)
            cell = f'<div>{val_str}</div>'
            if change:
                arrow, text, color = change
                cell += f'<div style="font-size:11px; color:{color};">{arrow} {text}</div>'
            html.append(f'<td style="text-align:right; padding:6px 10px; vertical-align:top;">{cell}</td>')
        html.append('</tr>')
    html.append('</tbody></table>')
    return ''.join(html)


# ─────────────────────────────────────────────────────────────
# 3) 규칙 기반 코멘트 문장 생성 (API 키 불필요)
# ─────────────────────────────────────────────────────────────
def manwon(v):
    return f"{round(v / 10000):,}"


def pct100(ratio):
    return f"{round(ratio * 100)}%"


def change_phrase(pct_change, noun, up_word='증가', down_word='감소'):
    if pct_change is None:
        return f"{noun} 변동 없음"
    if abs(pct_change) < 1:
        return f"{noun} 소폭 변동"
    direction = up_word if pct_change > 0 else down_word
    return f"{noun} {abs(pct_change):.0f}% {direction}"


def summary_sentence_1(m, label_curr):
    sales = m['SALES']['curr']
    roas = m['ROAS']['curr']
    sales_pct = m['SALES']['pct_change']
    trend = "증가" if (sales_pct or 0) >= 0 else "감소"
    return (f"{label_curr} 매출액 {manwon(sales)}만원 확보하며 ROAS {pct100(roas)} 기록, "
            f"전주 대비 매출 {abs(sales_pct or 0):.0f}% {trend}")


def summary_sentence_2(m, note):
    # 가장 크게 움직인 효율 지표 하나를 골라 설명
    candidates = [('CTR', 'CTR'), ('CPC', 'CPC'), ('CVR', 'CVR')]
    best = max(candidates, key=lambda kv: abs(m[kv[0]]['pct_change'] or 0))
    key, label = best
    phrase = change_phrase(m[key]['pct_change'], label)
    roas_prev, roas_curr = pct100(m['ROAS']['prev']), pct100(m['ROAS']['curr'])
    base = f"{phrase}하며 ROAS {roas_prev}→{roas_curr} 흐름"
    if note:
        base = f"{note} 영향으로 {base}"
    return base


def group_line(m, note, brand_label=None, num=None):
    imps_p = change_phrase(m['IMPS']['pct_change'], '노출')
    clicks_p = change_phrase(m['CLICK']['pct_change'], '유입')
    cpc_p = change_phrase(m['CPC']['pct_change'], 'CPC', up_word='상승', down_word='하락')
    sales = manwon(m['SALES']['curr'])
    roas_prev, roas_curr = pct100(m['ROAS']['prev']), pct100(m['ROAS']['curr'])
    prefix = f"{num}. {brand_label} : " if (brand_label and num) else ""
    line = f"{prefix}{imps_p}, {clicks_p}, {cpc_p}하며 주간 매출액 {sales}만원 확보, ROAS {roas_prev}→{roas_curr} 기록"
    if note:
        line += f" ({note})"
    return line


def extra_detail_line(m, note):
    trans_p = change_phrase(m['TRANS']['pct_change'], '전환수')
    cvr_p = change_phrase(m['CVR']['pct_change'], 'CVR', up_word='개선', down_word='하락')
    line = f"{trans_p}, {cvr_p}"
    if note:
        line += f" — {note}"
    return line


def generate_comments_rule_based(brief):
    """API 호출 없이, 증감률 크기에 따라 정해둔 문구 패턴을 조합해 셀별 문장을 만든다."""
    all_comments = {}
    for sheet_name, sheet_brief in brief.items():
        if sheet_name not in CELL_MAP:
            continue
        cell_texts = {}
        cell_map = CELL_MAP[sheet_name]

        for group, cells in cell_map.items():
            if group not in sheet_brief:
                continue
            data = sheet_brief[group]
            m = data['metrics']
            note = data.get('user_note', '')

            if group == 'TOTAL':
                lines = [summary_sentence_1(m, data['curr_week']), summary_sentence_2(m, note)]
            elif sheet_name == 'NAVER BS_Pet':
                num = BS_PET_BRAND_NUMBER.get(group)
                lines = [group_line(m, note, brand_label=group, num=num)]
            else:
                lines = [group_line(m, note), extra_detail_line(m, note)]
                if len(cells) >= 3:
                    lines.append(f"→ {group} 채널, 다음 주에도 동일 기조 유지 제안" if not note
                                 else f"→ {note} 관련 소재/운영 지속 제안")

            for cell, text in zip(cells, lines):
                cell_texts[cell] = text
            # 셀 수가 더 많으면 마지막 문장을 채워 넣음
            if len(cells) > len(lines):
                for c in cells[len(lines):]:
                    cell_texts[c] = lines[-1] if lines else ''

        all_comments[sheet_name] = cell_texts
    return all_comments


# ─────────────────────────────────────────────────────────────
# 4) Streamlit 화면
# ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="주간 리포트 코멘트 자동화", layout="wide")
st.title("📊 Report Automation Program")
st.caption("로우데이터 또는 완성된 리포트 파일을 올리면, 데이터 분석 → 특이사항 반영 → 코멘트 자동 작성까지 진행합니다. API 키가 필요 없습니다.")

mode = st.radio(
    "무엇을 업로드하시나요?",
    [
        "① 로우데이터 + 지난 리포트 파일 → 서식 그대로 이어서 업데이트 (추천)",
        "② 로우데이터만 → 새 리포트 생성 (서식 없이 데이터만)",
        "③ 이미 만들어진 리포트 파일 → 코멘트만 작성",
        "④ 성과 대시보드 보기 (주차별 추이 · 채널/브랜드 비교)",
    ],
)

file_bytes = None

if mode.startswith("①"):
    st.caption("지난 주까지 쓰던 실제 리포트 파일 + 오늘까지의 로우데이터를 올리면, 서식·다른 시트는 그대로 두고 완료된 주의 숫자만 채워 넣습니다.")
    col1, col2 = st.columns(2)
    with col1:
        prev_uploaded = st.file_uploader("지난 리포트 파일 업로드 (.xlsx)", type="xlsx", key="prev_uploader")
    with col2:
        raw_uploaded = st.file_uploader("로우데이터 파일 업로드 (.xlsx)", type="xlsx", key="raw_uploader_1")
    if prev_uploaded and raw_uploaded:
        with st.spinner("로우데이터를 분석해서 리포트에 반영하는 중..."):
            wb_updated, updated_cells = raw_to_report.update_existing_report(
                io.BytesIO(prev_uploaded.getvalue()), io.BytesIO(raw_uploaded.getvalue())
            )
            buf = io.BytesIO()
            wb_updated.save(buf)
            file_bytes = buf.getvalue()
        if updated_cells:
            weeks_done = sorted(set(u[2] for u in updated_cells))
            st.success(f"반영 완료 ({', '.join(weeks_done)}) — 서식은 원본 그대로입니다. 아래에서 이어서 특이사항을 입력하고 코멘트를 생성하세요.")
        else:
            st.warning("새로 반영할 완료된 주를 찾지 못했습니다. 로우데이터 기간을 확인해주세요.")

elif mode.startswith("②"):
    raw_uploaded = st.file_uploader("로우데이터 파일 업로드 (.xlsx)", type="xlsx", key="raw_uploader_2")
    st.caption("지원 매체/시트: GFA(Conf·Pet), NaverBS(Pet) — 컬럼: Date, Media, Device, 캠페인, 그룹, 노출, 클릭, 광고비, 구매완료수, 매출액. 서식 없이 데이터만 채워진 새 파일이 만들어집니다.")
    if raw_uploaded:
        with st.spinner("로우데이터를 분석해서 리포트를 생성하는 중..."):
            wb_generated = raw_to_report.build_report(io.BytesIO(raw_uploaded.getvalue()))
            buf = io.BytesIO()
            wb_generated.save(buf)
            file_bytes = buf.getvalue()
        st.success("리포트 생성 완료! 아래에서 이어서 특이사항을 입력하고 코멘트를 생성하세요.")

elif mode.startswith("④"):
    st.caption("숫자가 채워진 리포트 파일을 올리면 시트·채널/브랜드별 주차 추이와 비교 그래프를 보여줍니다. "
               "(엑셀에서 한 번 열어 저장한 파일을 올리면 더 정확합니다)")
    dash_uploaded = st.file_uploader("리포트 파일 업로드 (.xlsx)", type="xlsx", key="dash_uploader")

    if dash_uploaded:
        wb_dash = openpyxl.load_workbook(io.BytesIO(dash_uploaded.getvalue()), data_only=True)
        series = build_full_series(wb_dash)

        if not series:
            st.error("지원하는 시트(NAVER GFA_Conf / NAVER GFA_Pet / NAVER BS_Pet)에서 데이터를 찾지 못했습니다. "
                     "엑셀에서 파일을 한 번 열어 저장한 뒤 다시 업로드해보세요.")
        else:
            sheet_pick = st.selectbox("시트 선택", list(series.keys()), key="dash_sheet")
            groups_data = series[sheet_pick]

            # ── 주차별 추이 ──────────────────────────────────────
            st.subheader("📈 주차별 추이")
            default_trend_group = 'TOTAL' if 'TOTAL' in groups_data else list(groups_data.keys())[0]
            group_pick_trend = st.selectbox(
                "채널/브랜드 선택", list(groups_data.keys()),
                index=list(groups_data.keys()).index(default_trend_group), key="trend_group",
            )
            trend_source = groups_data[group_pick_trend]
            all_weeks_trend = [p['WEEK'] for p in trend_source]

            ym_labels = infer_year_month_labels(all_weeks_trend)
            week_to_ym = dict(zip(all_weeks_trend, ym_labels))
            unique_yms = sorted(set(ym_labels))

            col_a, col_b = st.columns(2)
            with col_a:
                start_ym = st.selectbox("시작월", unique_yms, index=0, key="trend_start_ym")
            with col_b:
                end_ym = st.selectbox("종료월", unique_yms, index=len(unique_yms) - 1, key="trend_end_ym")
            if start_ym > end_ym:
                start_ym, end_ym = end_ym, start_ym

            weeks_in_range = [w for w in all_weeks_trend if start_ym <= week_to_ym[w] <= end_ym]

            metric_pick_trend = st.multiselect(
                "지표 선택 (복수 선택 가능) — 주요지표는 막대, 보조지표는 선그래프로 표시됩니다",
                list(METRIC_LABELS.keys()), default=['SALES', 'ROAS'],
                format_func=lambda k: METRIC_LABELS[k], key="trend_metrics",
            )

            if weeks_in_range and metric_pick_trend:
                pts_by_week = {p['WEEK']: p for p in trend_source}
                series_dict = {m: [pts_by_week[w][m] for w in weeks_in_range] for m in metric_pick_trend}
                primary_sel = [m for m in metric_pick_trend if m in PRIMARY_METRICS]
                secondary_sel = [m for m in metric_pick_trend if m in SECONDARY_METRICS]
                render_combo_chart(
                    weeks_in_range, 'WEEK', series_dict, primary_sel, secondary_sel,
                    title=f"{group_pick_trend} 주차별 추이",
                )
            elif not weeks_in_range:
                st.info("선택한 기간에 데이터가 없습니다.")
            else:
                st.info("지표를 1개 이상 선택해주세요.")

            st.divider()

            # ── 채널/브랜드 비교 ──────────────────────────────────
            st.subheader("📊 채널/브랜드 비교")
            order_source = groups_data.get('TOTAL') or next(iter(groups_data.values()))
            all_weeks_cmp = [p['WEEK'] for p in order_source]
            weeks_picked = st.multiselect(
                "비교할 주차 선택 (복수 선택 가능)", all_weeks_cmp,
                default=[all_weeks_cmp[-1]], key="compare_weeks",
            )
            compare_groups = [g for g in groups_data if g != 'TOTAL']

            if not weeks_picked:
                st.info("주차를 1개 이상 선택해주세요.")
            elif len(weeks_picked) >= 2:
                metric_cmp = st.selectbox(
                    "지표 선택", list(METRIC_LABELS.keys()),
                    format_func=lambda k: METRIC_LABELS[k], key="compare_metric_multi",
                )
                fig_cmp = go.Figure()
                for w in weeks_picked:
                    y_vals = []
                    for g in compare_groups:
                        match = next((p for p in groups_data[g] if p['WEEK'] == w), None)
                        y_vals.append(match[metric_cmp] if match else None)
                    fig_cmp.add_bar(name=w, x=compare_groups, y=y_vals)  # 막대 이름 = 주차 레이블
                fig_cmp.update_layout(
                    barmode='group',
                    bargap=0.35,
                    bargroupgap=0.2,
                    xaxis=dict(title='채널/브랜드'),
                    yaxis=dict(title=METRIC_LABELS[metric_cmp]),
                    legend=dict(title='주차', orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                    margin=dict(t=70),
                )
                st.plotly_chart(fig_cmp, use_container_width=True)
            else:
                week_pick = weeks_picked[0]
                metric_pick_cmp = st.multiselect(
                    "지표 선택 (복수 선택 가능) — 주요지표는 막대, 보조지표는 선그래프로 표시됩니다",
                    list(METRIC_LABELS.keys()), default=['SALES', 'ROAS'],
                    format_func=lambda k: METRIC_LABELS[k], key="compare_metrics_single",
                )
                if metric_pick_cmp:
                    series_dict_cmp = {}
                    for m in metric_pick_cmp:
                        vals = []
                        for g in compare_groups:
                            match = next((p for p in groups_data[g] if p['WEEK'] == week_pick), None)
                            vals.append(match[m] if match else None)
                        series_dict_cmp[m] = vals
                    primary_sel_cmp = [m for m in metric_pick_cmp if m in PRIMARY_METRICS]
                    secondary_sel_cmp = [m for m in metric_pick_cmp if m in SECONDARY_METRICS]
                    render_combo_chart(
                        compare_groups, '채널/브랜드', series_dict_cmp, primary_sel_cmp, secondary_sel_cmp,
                        title=f"{week_pick} 채널/브랜드 비교",
                    )
                else:
                    st.info("지표를 1개 이상 선택해주세요.")

            if weeks_picked:
                st.markdown("#### 📋 선택 주차 상세 데이터")
                st.caption("각 셀 아래의 작은 숫자는 직전 주 대비 증감입니다. (▲ 빨강 = 증가, ▼ 파랑 = 감소, CTR·CVR은 %p 기준)")
                for w in weeks_picked:
                    st.markdown(f"**{w}**")
                    st.markdown(render_group_table_html(groups_data, w), unsafe_allow_html=True)
    st.stop()

else:
    uploaded = st.file_uploader("이번 주 리포트 파일 업로드 (.xlsx)", type="xlsx", key="report_uploader")
    if uploaded:
        file_bytes = uploaded.getvalue()

if file_bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    brief = build_brief(wb)

    if not brief:
        st.error("지원하는 시트(NAVER GFA_Conf / NAVER GFA_Pet / NAVER BS_Pet)에서 최근 완료된 주 데이터를 찾지 못했습니다. "
                 "엑셀에서 파일을 한 번 열어 저장한 뒤 다시 업로드해보세요 (수식 캐시 문제).")
    else:
        st.subheader("① 이번 주 특이사항 입력 (선택)")
        notes = {}
        cols = st.columns(len(brief))
        for col, (sheet_name, sheet_brief) in zip(cols, brief.items()):
            with col:
                st.markdown(f"**{sheet_name}**")
                for group in sheet_brief:
                    key = f"note_{sheet_name}_{group}"
                    notes[(sheet_name, group)] = st.text_input(group, key=key, placeholder="예: 15일 타임딜 운영")

        if st.button("② 코멘트 생성", type="primary"):
            for (sheet_name, group), memo in notes.items():
                if memo and memo.strip():
                    brief[sheet_name][group]['user_note'] = memo.strip()
            st.session_state['comments'] = generate_comments_rule_based(brief)
            st.session_state['file_bytes'] = file_bytes

if 'comments' in st.session_state:
    st.subheader("③ 생성된 코멘트 검수 (필요시 직접 수정)")
    edited = {}
    for sheet_name, cell_texts in st.session_state['comments'].items():
        st.markdown(f"**{sheet_name}**")
        edited[sheet_name] = {}
        for cell, text in cell_texts.items():
            edited[sheet_name][cell] = st.text_area(f"{sheet_name} · {cell}", value=text, key=f"edit_{sheet_name}_{cell}", height=80)

    if st.button("④ 엑셀에 반영하고 다운로드 생성", type="primary"):
        wb_out = openpyxl.load_workbook(io.BytesIO(st.session_state['file_bytes']), data_only=False)
        for sheet_name, cell_texts in edited.items():
            ws = wb_out[sheet_name]
            for cell, text in cell_texts.items():
                ws[cell] = text
        wb_out.calculation.fullCalcOnLoad = True
        buf = io.BytesIO()
        wb_out.save(buf)
        buf.seek(0)
        st.success("완료! 아래 버튼으로 다운로드하세요.")
        st.download_button(
            "📥 업데이트된 엑셀 다운로드",
            data=buf.getvalue(),
            file_name=f"report_updated_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
